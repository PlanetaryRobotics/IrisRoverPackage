"""
Module for looking up definitions from inside the latest Data Standards
(compiled from latest accessible FPrime).

Created: 10/29/2021
Last Update: 09/28/2022
"""
import pandas as pd
from styleframe import (  # type: ignore
    StyleFrame, Styler,
    utils as style_utils
)
from datetime import datetime
import openpyxl  # type: ignore
from openpyxl.utils import get_column_letter  # type: ignore
from openpyxl.styles import NamedStyle, Font, Border, Side  # type: ignore
from collections import OrderedDict
from typing import Optional
from termcolor import cprint, colored
from dataclasses import dataclass
from scripts.utils.trans_tools import *
import argparse

from IrisBackendv3.data_standards.module import Module, Argument, Command, Event, TelemetryChannel, EnumItem

parser = argparse.ArgumentParser(
    description='IRIS Lunar Rover — Data Standards Lookup — CLI')


def get_opts():
    """
    Return settings wrapped in argparse.

    Returns
    -------
    opts : settings wrapped in argparse.

    """
    parser.add_argument('-a', '--all', action='store_true',
                        help='Prints all the Data Standards')

    parser.add_argument('-l', '--list-modules', action='store_true',
                        help='Lists the names of modules registered in the Data Standards.')

    parser.add_argument('-t', '--telem-count', action='store_true',
                        help='Lists (and counts) all telemetry channels (useful for sizing hash buckets in FSW).')

    parser.add_argument('-m', '--module-name', type=str,
                        help='Print standards just for the module with the given name.')

    parser.add_argument('-f', '--print-to-file', type=str, default="",
                        help='Print standards to the excel file with the given name (ignoring xlsx extension).')

    opts = parser.parse_args()
    return opts


def print_modules_list() -> None:
    cprint("All Registered Modules:", 'magenta', 'on_grey', attrs=['bold'])
    for m in standards.modules.values():
        id_text = colored(f"0x{m.ID:04X}: ", 'blue')
        module_text = colored(f"{m}", 'green')
        print(f"\t {id_text} {module_text}")


# Styler args used for all table header cells:
STANDARD_EXCEL_HEADER_STYLE: Final[Dict[str, Any]] = dict(
    bg_color='00C00000',
    font_color='00FFFFFF',
    bold=True,
    horizontal_alignment=style_utils.horizontal_alignments.left
)
# Styler args used for all filler cells:
STANDARD_EXCEL_FILLER_STYLE: Final[Dict[str, Any]] = dict(
    bg_color='00D9D9D9',
)


@dataclass
class SheetContext:
    """Metadata where we're writing in a sheet.
    Field names match keys in `StyleFrame.to_excel`."""
    sheet_name: str
    startcol: int
    startrow: int


def add_styler_to_cell(cell: openpyxl.cell.Cell, styler: Styler):
    """Combine the given Styler with the style already present in the given
    cell."""
    cell.style = Styler.combine(
        Styler.from_openpyxl_style(cell),
        styler
    ).to_openpyxl_style()


def add_openpyxl_style_to_cell(cell: openpyxl.cell.Cell, style: NamedStyle):
    """Combine the given openpyxl NamedStyle with the style already present in
    the given cell."""
    add_styler_to_cell(cell, Styler.from_openpyxl_style(style))


def excel_width_to_sf(x: float) -> float:
    """For some reason, to achieve a certain width in excel, a different number
    has to be given to the StyleFrame. Can't find why. Encapsulating this
    behavior here in case it has to be changed."""
    return x//6+1


def set_sheet_column_width(
    width: int,
    writer: StyleFrame.ExcelWriter,
    sheet_context: SheetContext,
    col_inc: int = 0
) -> None:
    """Sets the width of column at the `sheet_context` head, in the sheet
    pointed to by `sheet_context` to have a width of `width` in Excel.
    Used for columns that aren't (necessarily) part of a table.
    If `col_inc` is given, the column adjusted will be at
    `sheet_context.startcol+col_inc`.
    Does not modify the `sheet_context` head.
    """
    sheet = writer.book[sheet_context.sheet_name]
    column_letter = get_column_letter(sheet_context.startcol+col_inc+1)
    sheet.column_dimensions[column_letter].width = excel_width_to_sf(width)


def style_table(table: pd.DataFrame) -> StyleFrame:
    """Helper function that applies standard styling to the given table and
    returns a StyleFrame."""
    # Convert index into an actual column:
    table.reset_index(inplace=True)
    # Create and style StyleFrame:
    sf = StyleFrame(table)
    # Add borders:
    sf = sf.apply_column_style(
        cols_to_style=table.columns,
        style_header=True,
        styler_obj=Styler(
            border_type=style_utils.borders.thin,
            bg_color="00FFFFFF",
            horizontal_alignment=style_utils.horizontal_alignments.left
        )
    )
    sf = sf.apply_headers_style(
        Styler(**STANDARD_EXCEL_HEADER_STYLE),
        style_index_header=True
    )
    return sf


def merge_right(
    sheet_name: str,
    row: int,
    col: int,
    width: int,
    writer: StyleFrame.ExcelWriter,
    post_merge_styler: Optional[Styler] = None
) -> None:
    """Merges (row,col) in sheet_name with cells to the right of it so it becomes
    `width` cells wide, using writer.
    `post_merge_styler` is applied after the merging since styling is lost.
    If none is given a simple default style is applied (aligns the text left).
    """
    # Merge cells if necessary:
    if width > 1:
        sheet = writer.book[sheet_name]
        # Merge:
        sheet.merge_cells(
            start_row=row+1,
            start_column=col+1,
            end_row=row+1,
            end_column=col+width
        )

        # Apply post-merge style:
        if post_merge_styler is None:
            # Use default style (align text left):
            post_merge_styler = Styler(
                horizontal_alignment=style_utils.horizontal_alignments.left
            )
        sheet = writer.book[sheet_name]
        cell = sheet.cell(row=row+1, column=col+1)  # 1-indexed
        cell.style = post_merge_styler.to_openpyxl_style()
        # writer.save()  # not part of public API, covered by .close()


def write_text(
    txt: str,
    row: int,
    col: int,
    writer: StyleFrame.ExcelWriter,
    sheet_context: SheetContext,
    styler: Optional[Styler] = None
) -> None:
    """ Writes the given txt at the given row and column with writer in the sheet pointed to by sheet_context.
    """
    sheet = writer.book[sheet_context.sheet_name]
    cell = sheet.cell(row=row+1, column=col+1)  # 1-indexed
    cell.value = txt
    if styler is not None:
        cell.style = styler.to_openpyxl_style()
    # writer.save()  # not part of public API, covered by .close()


def write_section_header(
    txt: str,
    writer: StyleFrame.ExcelWriter,
    sheet_context: SheetContext,
    width: int = 1,
    **kwargs
) -> None:
    """Writes the given section header at the current sheet_context head.
    If `width` is >1, the header will be merged with the cells to the right of
    it to become `width` cells wide.
    Any extra `kwargs` will be passed to the Styler."""
    # Merge cells if necessary (do this first so styles are applied after):
    merge_right(
        sheet_context.sheet_name,
        sheet_context.startrow,
        sheet_context.startcol,
        width,
        writer
    )

    write_text(
        txt,
        row=sheet_context.startrow,
        col=sheet_context.startcol,
        writer=writer,
        sheet_context=sheet_context,
        styler=Styler(
            bold=True,
            underline='single',
            border_type=style_utils.borders.hair,
            horizontal_alignment=style_utils.horizontal_alignments.left,
            **kwargs
        )
    )

    # Advance down:
    sheet_context.startrow += 1


def write_subsection_header(
    txt: str,
    writer: StyleFrame.ExcelWriter,
    sheet_context: SheetContext,
    width: int = 1,
    **kwargs
) -> None:
    """Writes the given subsection header at the current sheet_context head.
    If `width` is >1, the header will be merged with the cells to the right of
    it to become `width` cells wide.
    Any extra `kwargs` will be passed to the Styler."""
    # Merge cells if necessary (do this first so styles are applied after):
    merge_right(
        sheet_context.sheet_name,
        sheet_context.startrow,
        sheet_context.startcol,
        width,
        writer
    )

    write_text(
        txt,
        row=sheet_context.startrow,
        col=sheet_context.startcol,
        writer=writer,
        sheet_context=sheet_context,
        styler=Styler(
            bold=True,
            border_type=None,
            horizontal_alignment=style_utils.horizontal_alignments.left,
            **kwargs
        )
    )

    # Advance down:
    sheet_context.startrow += 1


def draw_section_border_from_end(
    n_rows: int,
    n_cols: int,
    writer: StyleFrame.ExcelWriter,
    sheet_context: SheetContext
) -> None:
    """Draws the border of a section box with the given num rows and cols that
    **ends** at the sheet_context head vertically
    (i.e. head is at the last row and first column of the box).
    Does not modify sheet_context head."""
    sheet = writer.book[sheet_context.sheet_name]
    border_side = Side(style='medium', color="000000")

    # Move context head vertically to beginning of box:
    sheet_context.startrow -= n_rows

    # Adds the given border style
    def _add_border_style(cell: openpyxl.cell.Cell, side: Side, edge: str) -> None:
        """Adds the given border style `side` to edge with the given name `edge`
        (e.g. 'top', 'left', etc.), keeping all other style attributes of that
        cell, including the styling of the other edges."""
        # Salt to add to NamedStyles to make unique:
        salt = str(hash(datetime.now()))
        old_style = cell  # `Cell` is a `StyleableObject` so it has all the requisite fields
        style = NamedStyle(
            f"sh_{sheet_context.sheet_name}_sect_{salt}_{edge}",
            font=old_style.font.__copy__(),
            fill=old_style.fill.__copy__(),
            # Use old border style but override the given edge with the given side:
            border=Border(**{
                **cell.border.__dict__,
                edge: side
            }),
            alignment=old_style.alignment.__copy__(),
            number_format=old_style.number_format,
            protection=old_style.protection.__copy__()
        )
        cell.style = style

    # Apply the border:
    for c in range(sheet_context.startcol, sheet_context.startcol+n_cols):
        # Top:
        cell = sheet.cell(row=sheet_context.startrow+1, column=c+1)
        _add_border_style(cell, border_side, 'top')
        # Bottom:
        cell = sheet.cell(row=sheet_context.startrow+n_rows, column=c+1)
        _add_border_style(cell, border_side, 'bottom')

    # left and right edges:
    for r in range(sheet_context.startrow, sheet_context.startrow+n_rows):
        # Left:
        cell = sheet.cell(row=r+1, column=sheet_context.startcol+1)
        _add_border_style(cell, border_side, 'left')
        # Right:
        cell = sheet.cell(row=r+1, column=sheet_context.startcol+n_cols)
        _add_border_style(cell, border_side, 'right')

    # Return context head to the end of the box:
    sheet_context.startrow += n_rows

    # Save changes to workbook before proceeding further:
    # writer.save()  # not part of public API, covered by .close()


def draw_section_fill_from_start(
    n_rows: int,
    n_cols: int,
    writer: StyleFrame.ExcelWriter,
    sheet_context: SheetContext
) -> None:
    """Draws the background fill of a section box with the given num rows and
    cols that  **starts** at the sheet_context head.
    Does not modify sheet_context head."""
    sheet = writer.book[sheet_context.sheet_name]
    bg_styler = Styler(
        bg_color="00D9D9D9",
        border_type=None
    )

    # Apply the background style:
    for r in range(sheet_context.startrow, sheet_context.startrow+n_rows):
        for c in range(sheet_context.startcol, sheet_context.startcol+n_cols):
            cell = sheet.cell(row=r+1, column=c+1)
            cell.style = bg_styler.to_openpyxl_style()

    # Save changes to workbook before proceeding further:
    # writer.save()  # not part of public API, covered by .close()


def write_dict_data(
    data: OrderedDict,
    writer: StyleFrame.ExcelWriter,
    sheet_context: SheetContext,
    data_width: int = 1
) -> None:
    """Writes the given OrderedDict of data at the sheet_context head.
    If `data_width` is >1, the value cell of each field will be merged with the
    cells to the right of it to become `data_width` cells wide.
    """
    for name, val in data.items():
        # Write the field name:
        write_text(
            str(name),
            row=sheet_context.startrow,
            col=sheet_context.startcol,
            writer=writer,
            sheet_context=sheet_context,
            styler=Styler(**STANDARD_EXCEL_HEADER_STYLE)
        )
        # Write the field val:
        write_text(
            str(val),
            row=sheet_context.startrow,
            col=sheet_context.startcol+1,
            writer=writer,
            sheet_context=sheet_context,
            styler=Styler(bg_color=style_utils.colors.white)  # default style
        )
        # Merge cells if necessary:
        merge_right(
            sheet_context.sheet_name,
            sheet_context.startrow,
            sheet_context.startcol+1,
            data_width,
            writer
        )

        # Advance down:
        sheet_context.startrow += 1


def write_table(
    table: StyleFrame,
    writer: StyleFrame.ExcelWriter,
    sheet_context: SheetContext,
    lateral_offset: int = 0,
    pre_gap_lines: int = 1
) -> None:
    """Writes the given style frame to the sheet_context head using writer.
    The whole table will be shifted right by `lateral_offset` cells if non-zero.
    A vertical gap of `pre_gap_lines` will be added before the table."""
    # Add pre-gap:
    sheet_context.startrow += pre_gap_lines
    # Adjust laterally:
    sheet_context.startcol += lateral_offset
    # Write table:
    table.to_excel(writer, **sheet_context.__dict__)
    # writer.save()  # testing shows we need to save here for some reason
    # not part of public API, covered by .close()
    # Advance rows:
    sheet_context.startrow += table.data_df.shape[0]+1  # +1 for header
    # Reset lateral adjustment:
    sheet_context.startcol -= lateral_offset


def write_titled_table(
    title: str,
    table: StyleFrame,
    writer: StyleFrame.ExcelWriter,
    sheet_context: SheetContext,
    lateral_offset: int = 0,
    pre_gap_lines: int = 1,
) -> None:
    """Writes the given title then the given style frame to the sheet_context
    head using writer.
    The whole table will be shifted right by `lateral_offset` cells if non-zero.
    A vertical gap of `post_gap_lines` will be added after the table."""
    # Add pre-gap:
    sheet_context.startrow += pre_gap_lines
    # Adjust laterally:
    sheet_context.startcol += lateral_offset
    write_subsection_header(title, writer, sheet_context,
                            width=table.data_df.shape[1],
                            # Style as a filler cell (bg_color, etc.):
                            **STANDARD_EXCEL_FILLER_STYLE)
    write_table(table, writer, sheet_context,
                lateral_offset=0, pre_gap_lines=0)
    # Reset lateral adjustment:
    sheet_context.startcol -= lateral_offset


def module_to_sheet_name(module: Module) -> str:
    """Get the Excel sheet name for the given module."""
    # Cap it at 31 chars (Excel limitation):
    return f"0x{module.ID:04X}_{module.name}"[:31]


def add_styler_to_table_column(
    table: StyleFrame,
    col_name: str,
    styler: Styler,
    style_header: bool = False
) -> None:
    """Adds the given styler to the column with the given name (combines with
    any existing style).
    The style for the given column is taken from the final cell in that column."""
    current_col_styles = table[col_name].style
    if isinstance(current_col_styles, pd.Series) and current_col_styles.size > 0:
        # col has currently applied styles. Merge with them:
        styler = Styler.combine(
            current_col_styles.iloc[-1],
            styler
        )

    table.apply_column_style(
        cols_to_style=[col_name],
        styler_obj=styler,
        style_header=style_header
    )


def make_arg_table(args: List[Argument]) -> StyleFrame:
    """Makes a StyleFrame from a list of args."""
    table = pd.DataFrame(
        columns=['Arg', 'Name', 'Type', 'Bytes (max)', 'Comment'],
        dtype=object
    ).set_index('Arg')

    for i, arg in enumerate(args):
        new_row = {
            'Name': arg.name,
            'Type': arg.datatype.name,
            'Bytes (max)': arg.datatype.num_octets,
            'Comment': "None"  # TODO: Add from gsw_metadata (once its used).
        }
        table.loc[i, [*new_row.keys()]] = [*new_row.values()]

    # Style the table:
    styled_table = style_table(table)
    # Set column widths:
    styled_table.set_column_width_dict({
        'Arg': excel_width_to_sf(80),
        'Name': excel_width_to_sf(150),
        'Type': excel_width_to_sf(110),
        'Bytes (max)': excel_width_to_sf(75),
        'Comment': excel_width_to_sf(200)
    })
    # Add column-specific styling:
    add_styler_to_table_column(styled_table, 'Arg', Styler(
        horizontal_alignment=style_utils.horizontal_alignments.center
    ))
    add_styler_to_table_column(styled_table, 'Bytes (max)', Styler(
        horizontal_alignment=style_utils.horizontal_alignments.right
    ))

    return styled_table


def make_enum_table(enum: List[EnumItem]) -> StyleFrame:
    """Makes a StyleFrame from the given list of EnumItems."""
    table = pd.DataFrame(
        columns=['Name', 'Value', 'Hex', 'Comment'],
        dtype=object
    ).set_index('Name')

    for ei in enum:
        new_row = {
            'Value': ei.value,
            'Hex': f"0x{ei.value:04X}",
            'Comment': ei.comment if ei.comment != "" else "None"
        }
        table.loc[ei.name, [*new_row.keys()]] = [*new_row.values()]

    styled_table = style_table(table)
    # Add column-specific styling:
    add_styler_to_table_column(styled_table, 'Value', Styler(
        horizontal_alignment=style_utils.horizontal_alignments.right
    ))
    add_styler_to_table_column(styled_table, 'Hex', Styler(
        horizontal_alignment=style_utils.horizontal_alignments.right
    ))

    return styled_table


def make_channel_table(module: Module) -> StyleFrame:
    """Makes a StyleFrame of the telemetry channels in the given module."""
    table = pd.DataFrame(
        columns=['Opcode', 'Name', 'Type', 'Bytes (max)', 'Comment'],
        dtype=object
    ).set_index('Opcode')

    for ch in module.telemetry.vals:
        opcode = (module.ID & 0xFF00) | ch.ID
        opcode_str = f"0x{opcode:04X} ({opcode:05d})"
        new_row = {
            'Name': ch.name,
            'Type': ch.datatype.name,
            'Bytes (max)': ch.datatype.num_octets,
            'Comment': "None"  # TODO: Add from gsw_metadata (once its used).
        }
        table.loc[opcode_str, [*new_row.keys()]] = [*new_row.values()]

    # Style the table:
    styled_table = style_table(table)
    # Set column widths:
    styled_table.set_column_width_dict({
        'Opcode': excel_width_to_sf(82),
        'Name': excel_width_to_sf(150),
        'Type': excel_width_to_sf(110),
        'Bytes (max)': excel_width_to_sf(75),
        'Comment': excel_width_to_sf(200)
    })
    # Add column-specific styling:
    add_styler_to_table_column(styled_table, 'Bytes (max)', Styler(
        horizontal_alignment=style_utils.horizontal_alignments.right
    ))

    return styled_table


def write_commands(
    module: Module,
    writer: StyleFrame.ExcelWriter,
    sheet_context: SheetContext
) -> None:
    """Writes the Commands of the given module using writer at the
    sheet_context head."""
    n_cols = 5  # Always has this many columns
    write_section_header("Commands", writer, sheet_context, width=n_cols)

    cprint(f"\t\t{len(module.commands)} Command(s)", 'green')

    # Set default column width for first column (if no tables written):
    set_sheet_column_width(75, writer, sheet_context)

    # Write all commands or "None" if none:
    if len(module.commands) == 0:
        write_subsection_header("None.", writer, sheet_context, width=n_cols)
    for cmd in module.commands.vals:
        has_args = len(cmd.args) > 0
        if has_args:
            # Add arg table:
            arg_table = make_arg_table(cmd.args)

        # Build enum (name, table) tuples (if applicable):
        enum_tables = []
        for i, arg in enumerate(cmd.args):
            if arg.is_enum:
                # Make table:
                et = make_enum_table(arg.enum)
                # Add table to collection:
                enum_tables.append((
                    f"Command Arg Enum Table: {arg.name} ({i})", et
                ))

        # Determine section dimensions:
        n_rows = (
            3  # header rows
            + (3+arg_table.data_df.shape[0] if has_args else 0)
            + sum(3+tbl.data_df.shape[0] for _, tbl in enum_tables)
        )

        # Draw the section box:
        draw_section_fill_from_start(n_rows, n_cols, writer, sheet_context)

        # Write section header:
        write_subsection_header(
            f"Command: {cmd.name}[{', '.join(a.name for a in cmd.args)}]",
            writer,
            sheet_context,
            width=n_cols,
            # Style as a filler cell (bg_color, etc.):
            **STANDARD_EXCEL_FILLER_STYLE
        )
        # Write section details:
        opcode = (module.ID & 0xFF00) | cmd.ID
        write_dict_data(
            OrderedDict([
                ('Opcode:', f"0x{opcode:04X} ({opcode:05d})"),
                ('Mnemonic:', cmd.mnemonic)
            ]),
            writer,
            sheet_context,
            data_width=n_cols-1
        )

        # Write arg_table:
        if has_args:
            write_titled_table(
                "Command Args Table:", arg_table,
                writer, sheet_context
            )
        # Write any enums:
        for name, table in enum_tables:
            write_titled_table(
                name, table, writer, sheet_context,
                # indent table (so ends line up with arg table):
                lateral_offset=1
            )

        # Draw the section box border:
        draw_section_border_from_end(n_rows, n_cols, writer, sheet_context)

        # Add a vertical gap after command block:
        sheet_context.startrow += 2

    # Advance head to the right-most edge of the section
    # (after all commands written):
    sheet_context.startcol += n_cols-1


def write_telem(
    module: Module,
    writer: StyleFrame.ExcelWriter,
    sheet_context: SheetContext
) -> None:
    """Writes the Telemetry Channels of the given module using writer at the
    sheet_context head."""
    n_cols = 5  # Always has this many columns
    write_section_header("Telemetry Channels", writer,
                         sheet_context, width=n_cols)

    cprint(f"\t\t{len(module.telemetry)} Telemetry Channel(s)", 'green')

    # Set default column width for first column (if no tables written):
    set_sheet_column_width(75, writer, sheet_context)

    # Write all commands or "None" if none:
    if len(module.telemetry) == 0:
        write_subsection_header("None.", writer, sheet_context, width=n_cols)
    else:
        channel_table = make_channel_table(module)

        # Build enum (name, table) tuples (if applicable):
        enum_tables = []
        for ch in module.telemetry.vals:
            if ch.is_enum:
                # Make table:
                et = make_enum_table(ch.enum)
                # Add table to collection:
                opcode = (module.ID & 0xFF00) | ch.ID
                enum_tables.append((
                    f"Telemetry Channel Enum Table: {ch.name} (0x{opcode:04X})",
                    et
                ))

        # Determine section dimensions:
        n_rows = (
            0  # no header rows
            + 2+channel_table.data_df.shape[0]
            + sum(3+tbl.data_df.shape[0] for _, tbl in enum_tables)
        )

        # Draw the section box:
        draw_section_fill_from_start(n_rows, n_cols, writer, sheet_context)

        # Write channels table:
        write_titled_table(
            "Telemetry Channels Table:", channel_table,
            writer, sheet_context,
            pre_gap_lines=0  # first thing in the section block, so no pre-gap
        )
        # Write any enums:
        for name, table in enum_tables:
            write_titled_table(
                name, table, writer, sheet_context,
                # indent table (so ends line up with arg table):
                lateral_offset=1
            )

        # Draw the section box border:
        draw_section_border_from_end(n_rows, n_cols, writer, sheet_context)

    # Advance head to the right-most edge of the section
    # (after all channels written):
    sheet_context.startcol += n_cols-1


# Max event Severity Level:
MAX_EVENT_SEVERITY_LEVEL_VALUE: Final[int] = max(
    e.value for e in Event.SeverityLevel)


def write_events(
    module: Module,
    writer: StyleFrame.ExcelWriter,
    sheet_context: SheetContext
) -> None:
    """Writes the Events of the given module using writer at the
    sheet_context head."""
    n_cols = 5  # Always has this many columns
    write_section_header("Events", writer, sheet_context, width=n_cols)

    cprint(f"\t\t{len(module.events)} Event(s)", 'green')

    # Set default column width for first column (if no tables written):
    set_sheet_column_width(80, writer, sheet_context)

    # Write all commands or "None" if none:
    if len(module.events) == 0:
        write_subsection_header("None.", writer, sheet_context, width=n_cols)
    for evt in module.events.vals:
        has_args = len(evt.args) > 0
        if has_args:
            # Add arg table:
            arg_table = make_arg_table(evt.args)

        # Build enum (name, table) tuples (if applicable):
        enum_tables = []
        for i, arg in enumerate(evt.args):
            if arg.is_enum:
                # Make table:
                et = make_enum_table(arg.enum)
                # Add table to collection:
                enum_tables.append((
                    f"Event Arg Enum Table: {arg.name} ({i})", et
                ))

        # Determine section dimensions:
        n_rows = (
            4  # header rows
            + (3+arg_table.data_df.shape[0] if has_args else 0)
            + sum(3+tbl.data_df.shape[0] for _, tbl in enum_tables)
        )

        # Draw the section box:
        draw_section_fill_from_start(n_rows, n_cols, writer, sheet_context)

        # Write section header:
        write_subsection_header(
            f"Event: {evt.name}[{', '.join(a.name for a in evt.args)}]",
            writer,
            sheet_context,
            width=n_cols,
            # Style as a filler cell (bg_color, etc.):
            **STANDARD_EXCEL_FILLER_STYLE
        )
        # Write section details:
        opcode = (module.ID & 0xFF00) | evt.ID
        write_dict_data(
            OrderedDict([
                ('Opcode:', f"0x{opcode:04X} ({opcode:05d})"),
                ('Severity:',
                 f"{evt.severity.name} ({evt.severity.value}/{MAX_EVENT_SEVERITY_LEVEL_VALUE}): {evt.severity.description}"),
                ('Format String:', evt.format_string)
            ]),
            writer,
            sheet_context,
            data_width=n_cols-1
        )

        # Write arg_table:
        if has_args:
            write_titled_table(
                "Event Args Table:", arg_table,
                writer, sheet_context
            )
        # Write any enums:
        for name, table in enum_tables:
            write_titled_table(
                name, table, writer, sheet_context,
                # indent table (so ends line up with arg table):
                lateral_offset=1
            )

        # Draw the section box border:
        draw_section_border_from_end(n_rows, n_cols, writer, sheet_context)

        # Add a vertical gap after command block:
        sheet_context.startrow += 2

    # Advance head to the right-most edge of the section
    # (after all events written):
    sheet_context.startcol += n_cols-1


def write_module(module: Module, writer: StyleFrame.ExcelWriter) -> None:
    # Start a new sheet:
    sheet_context = SheetContext(
        sheet_name=module_to_sheet_name(module),
        # Track where to write the next DataFrame:
        startcol=0,
        startrow=0
    )
    cprint(
        f"\tMaking sheet {sheet_context.sheet_name} . . .",
        'magenta', attrs=['bold']
    )

    # Add the sheet:
    sheet = writer.book.create_sheet(sheet_context.sheet_name)
    sheet.title = sheet_context.sheet_name
    writer.sheets.update({ws.title: ws for ws in writer.book.worksheets})

    # Add module header:
    write_section_header(
        (
            f"Module: {module.name} (0x{module.ID:04X} = {module.ID}): "
            f"{len(module.commands)} Command(s), "
            f"{len(module.telemetry)} Telemetry Channel(s), "
            f"{len(module.events)} Event(s)."
        ),
        writer,
        sheet_context,
        width=17
    )

    # Add a vertical gap:
    sheet_context.startrow += 1

    # Start all data sections from here:
    first_data_row = sheet_context.startrow

    # Write commands:
    write_commands(module, writer, sheet_context)
    # writer.save()  # not part of public API, covered by .close()

    # Shift over, add a separator column, shift over:
    sheet_context.startcol += 1
    set_sheet_column_width(8, writer, sheet_context)
    sheet_context.startcol += 1

    # Write Telemetry Channels:
    sheet_context.startrow = first_data_row  # Reset sheet_context to the top
    write_telem(module, writer, sheet_context)
    # writer.save()  # not part of public API, covered by .close()

    # Shift over, add a separator column, shift over:
    sheet_context.startcol += 1
    set_sheet_column_width(8, writer, sheet_context)
    sheet_context.startcol += 1

    # Write Telemetry Channels:
    sheet_context.startrow = first_data_row  # Reset sheet_context to the top
    write_events(module, writer, sheet_context)
    # writer.save()  # not part of public API, covered by .close()


def print_to_file(file_name: str) -> None:
    """Prints all datastandards data to an excel file."""
    # Create the writer:
    path = f"{file_name}.xlsx"
    cprint(f"Writing to `./{path}` . . .",  'blue')
    book = openpyxl.Workbook()
    with StyleFrame.ExcelWriter(path) as writer:
        writer.book = book
        for module in standards.modules.vals:
            write_module(module, writer)
        # writer.save()  # not part of public API, covered by .close()


def print_lookup(module_to_lookup: Optional[str] = None) -> None:
    if module_to_lookup is None:
        standards.print_overview()
    else:
        def module(x): return cprint(
            f"\n\t{x}", 'magenta', 'on_grey', attrs=['bold'])

        def header(x): return cprint(f"\n\t\t{x}", 'grey', 'on_white')
        def command(i, x): return cprint(f"\n\t\t\t{i}.\t{x}", 'green')
        def telemetry(i, x): return cprint(f"\n\t\t\t{i}.\t{x}", 'red')
        def event(i, x): return cprint(f"\n\t\t\t{i}.\t{x}", 'blue')

        def p_arg(a): return cprint(
            f"\n\t\t\t\t\tAvailable values for `{a.name}`:", 'cyan')

        def p_enum(x): return cprint(
            f"\n\t\t\t\t\t\t'{x.name}' or {x.value} or {hex(x.value)}", 'magenta')

        print("Data Standards Overview: [")
        m = standards.modules[module_to_lookup]
        module(m)
        header('Commands:')
        for i, c in enumerate(m.commands.vals):
            command(i, c)
            for arg in c.args:
                if arg.is_enum:
                    p_arg(arg)
                    for e in arg.enum:
                        p_enum(e)
        header('Telemetry:')
        for i, t in enumerate(m.telemetry.vals):
            telemetry(i, t)
            if t.is_enum:
                for e in t.enum:
                    p_enum(e)
        header('Events:')
        for i, ev in enumerate(m.events.vals):
            event(i, ev)
            for arg in ev.args:
                if arg.is_enum:
                    p_arg(arg)
                    for e in arg.enum:
                        p_enum(e)
        print('\n]')


def all_telem_channels() -> None:
    # Useful for determining number of hash buckets needed in TLMCHAN_HASH_BUCKETS in Svc/TlmChan/TlmChanImplCfg.hpp in the FPrime FSW.
    def header(x): return cprint(f"\n\t{x}", 'grey', 'on_white')

    header('All Telemetry Channels:')
    tlm_count = 0
    for ((m_id, m_name), m) in standards.modules.fast_items():
        module_tlm_count = 0
        for idx, t in enumerate(m.telemetry.vals):
            cprint(f"\n\t\t{m_name}: \t{idx}.\t{t}", 'red', 'on_white')
            module_tlm_count += 1
            tlm_count += 1
        print(f"Telemetry Channels in {m_name}: {module_tlm_count}")

    header(f"There are {tlm_count} Total Telemetry Channels in the System.")


if __name__ == '__main__':
    opts = get_opts()

    parser.print_help()

    if opts.print_to_file != "":
        print_to_file(opts.print_to_file)

    if opts.all:
        print_lookup()

    if opts.module_name:
        print_lookup(opts.module_name)

    if opts.list_modules:
        print_modules_list()

    if opts.telem_count:
        all_telem_channels()
